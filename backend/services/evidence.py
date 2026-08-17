from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from threading import RLock

import fitz

from document_normalizer import CanonicalDocument


def _normalized(value: str) -> str:
    return " ".join(re.findall(r"[a-z0-9]+", value.casefold()))


class PreviewService:
    """Create page-stable DOCX previews and cache block-to-page mappings."""

    def __init__(self, root: Path) -> None:
        self.root = root
        self.root.mkdir(parents=True, exist_ok=True)
        self._lock = RLock()

    def ensure(self, original: Path, document: CanonicalDocument) -> tuple[Path, dict[str, int]]:
        pdf_path = self.root / f"{document.file_hash}.pdf"
        map_path = self.root / f"{document.file_hash}.pages.json"
        with self._lock:
            if not pdf_path.is_file():
                self._convert_with_word(original, pdf_path)
            if map_path.is_file():
                return pdf_path, json.loads(map_path.read_text(encoding="utf-8"))
            pages = self._page_text(pdf_path)
            mapping = self._map_blocks(document, pages)
            map_path.write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
            return pdf_path, mapping

    @staticmethod
    def _convert_with_word(original: Path, destination: Path) -> None:
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            raise RuntimeError("DOCX preview requires Microsoft Word or LibreOffice.") from exc
        pythoncom.CoInitialize()
        word = None
        document = None
        try:
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            document = word.Documents.Open(str(original.resolve()), ReadOnly=True)
            document.ExportAsFixedFormat(str(destination.resolve()), 17)
        except Exception as exc:
            destination.unlink(missing_ok=True)
            raise RuntimeError(f"Microsoft Word could not render the DOCX preview: {exc}") from exc
        finally:
            if document is not None:
                document.Close(False)
            if word is not None:
                word.Quit()
            pythoncom.CoUninitialize()

    @staticmethod
    def _page_text(pdf_path: Path) -> list[str]:
        with fitz.open(pdf_path) as pdf:
            return [_normalized(page.get_text("text")) for page in pdf]

    @staticmethod
    def _map_blocks(document: CanonicalDocument, pages: list[str]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        cursor = 0
        for block in document.blocks:
            needle = _normalized(block.text)[:180]
            if not needle:
                continue
            candidates = list(range(cursor, len(pages))) + list(range(0, cursor))
            match = next((index for index in candidates if needle[:80] in pages[index]), None)
            if match is None:
                tokens = set(needle.split()[:24])
                scores = [len(tokens & set(page.split())) for page in pages]
                match = max(range(len(scores)), key=scores.__getitem__) if scores and max(scores) >= 2 else cursor
            mapping[block.block_id] = match + 1
            cursor = match
        return mapping


def read_tabular_evidence(path: Path, sheet: str | None = None) -> dict[str, object]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            if sheet not in workbook.sheetnames:
                sheet = workbook.sheetnames[0]
            worksheet = workbook[sheet]
            rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]
            return {"kind": "xlsx", "sheet": sheet, "sheets": workbook.sheetnames,
                    "rows": rows, "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column}
        finally:
            workbook.close()
    if suffix == ".csv":
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
        return {"kind": "csv", "sheet": None, "sheets": [], "rows": rows,
                "max_row": len(rows), "max_column": max(map(len, rows), default=0)}
    raise ValueError("Internal table viewer supports XLSX and CSV files only.")
