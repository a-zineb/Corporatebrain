"""Deterministic, in-memory normalization for supported enterprise documents.

This module does not index data or mutate Chroma.  Every format converges to
``CanonicalDocument`` and ``CanonicalBlock`` while preserving source identity.
"""

from __future__ import annotations

import csv
import io
import re
import tempfile
import zipfile
from dataclasses import asdict, dataclass, field
from hashlib import sha256
from pathlib import Path, PurePosixPath
from typing import Any, Callable

from structured_ingestion import extract_docx_blocks
from document_engine import LogicalTable, read_docx_structure, read_pdf_structure


SUPPORTED_EXTENSIONS = frozenset({".docx", ".doc", ".pdf", ".xlsx", ".csv", ".zip"})
CANONICAL_SCHEMA_VERSION = 3
_SECRET_LABEL = re.compile(r"(?:password|passwd|token|api[_ -]?key|secret|private\s+key)", re.I)
_SECRET_ASSIGNMENT = re.compile(
    r"(?i)(\b(?:password|passwd|token|api[_ -]?key|secret|private\s+key)\s*[:=]\s*)"
    r"([^|;,\n]+)"
)


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _redact_value(label: str, value: Any) -> str:
    return "[REDACTED]" if _SECRET_LABEL.search(label) else _clean(value)


def _redact_text(text: Any) -> str:
    clean = _clean(text)
    return _SECRET_ASSIGNMENT.sub(lambda match: match.group(1) + "[REDACTED]", clean)


@dataclass(frozen=True)
class CanonicalBlock:
    block_id: str
    text: str
    block_type: str
    source_file: str
    file_hash: str
    section: str | None = None
    subsection: str | None = None
    page: int | None = None
    sheet: str | None = None
    table_index: int | None = None
    row_index: int | None = None
    paragraph_index: int | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class CanonicalDocument:
    document_id: str
    file_hash: str
    source_file: str
    file_type: str
    blocks: tuple[CanonicalBlock, ...]
    logical_tables: tuple[LogicalTable, ...] = ()
    sections: tuple[tuple[str, ...], ...] = ()
    warnings: tuple[str, ...] = ()
    metadata: dict[str, Any] = field(default_factory=dict)

    def canonical_text(self) -> str:
        lines = ["[DOCUMENT]", f"source_file = {self.source_file}", f"file_hash = {self.file_hash}"]
        for table in self.logical_tables:
            path = " > ".join(table.section_path)
            lines.extend(["", f'[LOGICAL_TABLE id="{table.table_id}" shape="{table.shape}" section="{path}"]'])
            values = table.values()
            for row_index, row in enumerate(values):
                positional = " | ".join(f"c{column_index}={value}" for column_index, value in enumerate(row))
                lines.append(f"row={row_index} | {positional}")
        previous_section: str | None = None
        for block in self.blocks:
            if block.section and block.section != previous_section:
                lines.extend(["", "[SECTION]", block.section])
                previous_section = block.section
            label = {
                "heading": "HEADING",
                "key_value": "KEY_VALUE",
                "table_row": "TABLE_ROW",
                "paragraph": "PARAGRAPH",
            }.get(block.block_type, block.block_type.upper())
            lines.extend(["", f"[{label}]", block.text])
        if self.warnings:
            lines.extend(["", "[WARNINGS]", *self.warnings])
        return "\n".join(lines) + "\n"

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["blocks"] = [asdict(block) for block in self.blocks]
        data["logical_tables"] = [asdict(table) for table in self.logical_tables]
        return data


def _make_block(*, ordinal: int, text: str, block_type: str, source_file: str,
                file_hash: str, **location: Any) -> CanonicalBlock:
    clean = _redact_text(text)
    identity = "\0".join((file_hash, source_file, str(ordinal), block_type, clean))
    return CanonicalBlock(
        block_id=sha256(identity.encode("utf-8")).hexdigest(), text=clean,
        block_type=block_type, source_file=source_file, file_hash=file_hash, **location,
    )


def _docx_blocks(data: bytes, source_file: str, file_hash: str) -> list[CanonicalBlock]:
    result: list[CanonicalBlock] = []
    paragraph_index = 0
    extracted = extract_docx_blocks(data, source_file)
    # The legacy extractor treats any wide table beginning with "System" as a
    # vertical matrix.  A horizontal one-record table then appears as several
    # ``BI = value`` blocks. Reconstruct that row from retained column headers.
    repaired: list[Any] = []
    index = 0
    while index < len(extracted):
        old = extracted[index]
        if (old.metadata.get("normalization_strategy") == "column_records"
                and old.metadata.get("logical_table_shape") != "MATRIX"):
            group = [old]
            cursor = index + 1
            while cursor < len(extracted) and extracted[cursor].table_index == old.table_index and extracted[cursor].metadata.get("normalization_strategy") == "column_records":
                group.append(extracted[cursor])
                cursor += 1
            entities = [_clean(item.text.split("=", 1)[0]) for item in group if "=" in item.text]
            if group and len(set(entities)) == 1 and entities[0]:
                entity = entities[0]
                pairs = [f"System = {entity}"]
                for item in group:
                    header = _clean(item.metadata.get("column_header"))
                    value = _clean(item.text.split("=", 1)[1])
                    pairs.append(f"{header or 'Value'} = {_redact_value(header, value)}")
                old.text = " | ".join(pairs)
                old.metadata = {**old.metadata, "normalization_strategy": "repaired_row_records"}
                repaired.append(old)
                index = cursor
                continue
        repaired.append(old)
        index += 1
    for ordinal, old in enumerate(repaired):
        block_type = old.block_type
        if old.metadata.get("normalization_strategy") == "key_value":
            block_type = "key_value"
        result.append(_make_block(
            ordinal=ordinal, text=old.text, block_type=block_type,
            source_file=source_file, file_hash=file_hash, section=old.section,
            table_index=old.table_index, row_index=old.row_index,
            paragraph_index=paragraph_index if old.block_type in {"paragraph", "heading"} else None,
            metadata={**old.metadata, "location": old.location},
        ))
        if old.block_type in {"paragraph", "heading"}:
            paragraph_index += 1
    return result


def _pdf_blocks(data: bytes, source_file: str, file_hash: str) -> tuple[list[CanonicalBlock], list[str], tuple[LogicalTable, ...]]:
    blocks: list[CanonicalBlock] = []
    structure = read_pdf_structure(data)
    ordinal = 0
    for table in structure.logical_tables:
        rows = table.values()
        if len(rows) >= 3 and len(rows[0]) >= 3:
            top = [value for value in rows[0][1:] if value]
            if top and len(set(_norm.casefold() for _norm in top)) == 1 and sum(bool(v) for v in rows[1][1:]) >= 2:
                rows = [rows[1], *rows[2:]]
        table_blocks = _rows_to_blocks(
            rows, source_file=source_file, file_hash=file_hash, sheet=None,
            ordinal_start=ordinal, allow_two_column_key_values=True,
        )
        for table_block in table_blocks:
            payload = asdict(table_block)
            payload["page"] = table.page
            payload["table_index"] = table.source_order
            payload["section"] = table.section
            payload["metadata"] = {
                **payload["metadata"], "section_path": list(table.section_path),
                "logical_table_shape": table.shape,
                "page_end": table.metadata.get("page_end", table.page),
                "cross_page": bool(table.metadata.get("cross_page")),
            }
            blocks.append(CanonicalBlock(**payload))
        if table.shape == "KEY_VALUE":
            for row_index, row in enumerate(rows):
                if len(row) < 2:
                    continue
                parent, value = _clean(row[0]), _clean(row[1])
                # Recover labelled records flattened into one PDF cell.
                subfields = list(re.finditer(
                    r"(?i)\b(IP|Login|Password)\s*[:=]\s*(.*?)(?=\s+\b(?:IP|Login|Password)\s*[:=]|$)", value
                ))
                for sub_index, match in enumerate(subfields):
                    label, subvalue = match.group(1), match.group(2).strip()
                    blocks.append(_make_block(
                        ordinal=ordinal + len(table_blocks) + row_index * 10 + sub_index,
                        text=f"{parent} | {label} = {_redact_value(label, subvalue)}",
                        block_type="key_value", source_file=source_file, file_hash=file_hash,
                        page=table.page, section=table.section, table_index=table.source_order,
                        row_index=row_index, metadata={"parent_field": parent, "subfield": label,
                                                       "section_path": list(table.section_path)},
                    ))
                if "directory" in parent.casefold():
                    alternatives = [part.strip(" _") for part in re.split(r"\s+_+(?:\s+_+)*\s+", value)
                                    if part.strip(" _").startswith("/")]
                    for alternative_index, alternative in enumerate(alternatives):
                        blocks.append(_make_block(
                            ordinal=ordinal + len(table_blocks) + row_index * 10 + 5 + alternative_index,
                            text=f"{parent} = {alternative}", block_type="key_value",
                            source_file=source_file, file_hash=file_hash, page=table.page,
                            section=table.section, table_index=table.source_order, row_index=row_index,
                            metadata={"alternative_value": True, "section_path": list(table.section_path)},
                        ))
        if table.shape == "GLOSSARY" and rows:
            for row_index, row in enumerate(rows[1:], start=1):
                term = _clean(row[0]) if row else ""
                value = _clean(row[1]) if len(row) > 1 else ""
                if term and not value:
                    blocks.append(_make_block(
                        ordinal=ordinal + len(table_blocks) + row_index,
                        text=f"{term} = [NO_VALUE]", block_type="key_value",
                        source_file=source_file, file_hash=file_hash, page=table.page,
                        section=table.section, table_index=table.source_order, row_index=row_index,
                        metadata={"explicit_empty_value": True, "logical_table_shape": "GLOSSARY",
                                  "section_path": list(table.section_path)},
                    ))
        ordinal += len(table_blocks)
    for paragraph_index, item in enumerate(structure.blocks):
        blocks.append(_make_block(
            ordinal=ordinal, text=item.text, block_type=item.block_type, source_file=source_file,
            file_hash=file_hash, page=item.page, paragraph_index=paragraph_index,
            section=item.section_path[-1] if item.section_path else None,
            metadata={"bbox": [round(value, 3) for value in item.bbox], "font_size": item.font_size,
                      "bold": item.bold, "section_path": list(item.section_path)},
        ))
        ordinal += 1
    warnings = list(structure.warnings)
    if not blocks:
        warnings.append("No embedded text was extracted; OCR may be required.")
    return blocks, warnings, structure.logical_tables


def _rows_to_blocks(rows: list[list[Any]], *, source_file: str, file_hash: str,
                    sheet: str | None, ordinal_start: int = 0,
                    allow_two_column_key_values: bool = False) -> list[CanonicalBlock]:
    nonempty = [row for row in rows if any(_clean(value) for value in row)]
    if not nonempty:
        return []
    width = max(len(row) for row in nonempty)
    padded = [list(row) + [None] * (width - len(row)) for row in nonempty]
    first = [_clean(value) for value in padded[0]]
    first_column = [_clean(row[0]) for row in padded]
    matrix_labels = {"protocol", "host", "hostname", "username", "password", "filedirectory", "directory", "port"}
    is_matrix = width >= 3 and any(label.casefold() in matrix_labels for label in first_column[1:])
    valid_headers = len([value for value in first if value]) >= 2 and len(set(v.casefold() for v in first if v)) == len([v for v in first if v])
    blocks: list[CanonicalBlock] = []
    if is_matrix:
        for column_index in range(1, width):
            pairs = []
            if first[column_index]:
                pairs.append(f"{first[0] or 'Column 1'} = {_redact_value(first[0], first[column_index])}")
            for row in padded[1:]:
                label, value = _clean(row[0]), row[column_index]
                if label and _clean(value):
                    pairs.append(f"{label} = {_redact_value(label, value)}")
            if pairs:
                blocks.append(_make_block(
                    ordinal=ordinal_start + len(blocks), text=" | ".join(pairs), block_type="table_row",
                    source_file=source_file, file_hash=file_hash, sheet=sheet, row_index=column_index,
                    metadata={"column_header": first[column_index], "normalization_strategy": "column_records"},
                ))
    elif allow_two_column_key_values and width == 2 and len(padded) > 1 and all(_clean(row[0]) for row in padded):
        for source_row, row in enumerate(padded, start=1):
            label, value = _clean(row[0]), row[1]
            if _clean(value):
                blocks.append(_make_block(
                    ordinal=ordinal_start + len(blocks), text=f"{label} = {_redact_value(label, value)}",
                    block_type="key_value", source_file=source_file, file_hash=file_hash,
                    sheet=sheet, row_index=source_row,
                    metadata={"normalization_strategy": "key_value"},
                ))
    elif valid_headers:
        for source_row, row in enumerate(padded[1:], start=2):
            pairs = [f"{first[index] or f'Column {index + 1}'} = {_redact_value(first[index], value)}"
                     for index, value in enumerate(row) if _clean(value)]
            if pairs:
                blocks.append(_make_block(
                    ordinal=ordinal_start + len(blocks), text=" | ".join(pairs), block_type="table_row",
                    source_file=source_file, file_hash=file_hash, sheet=sheet, row_index=source_row,
                    metadata={"column_headers": first, "normalization_strategy": "row_records"},
                ))
    else:
        for source_row, row in enumerate(padded, start=1):
            values = [f"Column {index + 1} = {_clean(value)}" for index, value in enumerate(row) if _clean(value)]
            if values:
                blocks.append(_make_block(
                    ordinal=ordinal_start + len(blocks), text=" | ".join(values), block_type="table_row",
                    source_file=source_file, file_hash=file_hash, sheet=sheet, row_index=source_row,
                    metadata={"normalization_strategy": "headerless_rows"},
                ))
    return blocks


def _xlsx_blocks(data: bytes, source_file: str, file_hash: str) -> list[CanonicalBlock]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    blocks: list[CanonicalBlock] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            blocks.extend(_rows_to_blocks(rows, source_file=source_file, file_hash=file_hash,
                                          sheet=worksheet.title, ordinal_start=len(blocks),
                                          allow_two_column_key_values=True))
    finally:
        workbook.close()
    return blocks


def _decode_csv(data: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", data, 0, len(data), "unsupported encoding")


def _csv_blocks(data: bytes, source_file: str, file_hash: str) -> tuple[list[CanonicalBlock], list[str]]:
    text, encoding = _decode_csv(data)
    warnings: list[str] = []
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        warnings.append("CSV dialect could not be detected; comma delimiter was used.")
    rows = list(csv.reader(io.StringIO(text), dialect))
    blocks = _rows_to_blocks(rows, source_file=source_file, file_hash=file_hash, sheet=None)
    for index, block in enumerate(blocks):
        metadata = {**block.metadata, "encoding": encoding}
        blocks[index] = CanonicalBlock(**{**asdict(block), "metadata": metadata})
    return blocks, warnings


def _safe_zip_name(name: str) -> bool:
    path = PurePosixPath(name.replace("\\", "/"))
    return bool(path.parts) and not path.is_absolute() and ".." not in path.parts


def _zip_blocks(data: bytes, source_file: str, file_hash: str) -> tuple[list[CanonicalBlock], list[str]]:
    blocks: list[CanonicalBlock] = []
    warnings: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        for info in sorted(archive.infolist(), key=lambda item: item.filename):
            suffix = Path(info.filename).suffix.lower()
            if info.is_dir() or not _safe_zip_name(info.filename) or suffix not in SUPPORTED_EXTENSIONS or suffix == ".zip":
                if not info.is_dir() and (not _safe_zip_name(info.filename) or suffix not in SUPPORTED_EXTENSIONS):
                    warnings.append(f"Ignored unsafe or unsupported ZIP entry: {info.filename}")
                continue
            inner_source = f"{source_file}::{info.filename}"
            try:
                inner = normalize_document(archive.read(info), inner_source)
            except Exception as exc:
                warnings.append(f"Could not normalize ZIP entry {info.filename}: {type(exc).__name__}")
                continue
            for inner_block in inner.blocks:
                blocks.append(_make_block(
                    ordinal=len(blocks), text=inner_block.text, block_type=inner_block.block_type,
                    source_file=inner_source, file_hash=file_hash, section=inner_block.section,
                    subsection=inner_block.subsection, page=inner_block.page, sheet=inner_block.sheet,
                    table_index=inner_block.table_index, row_index=inner_block.row_index,
                    paragraph_index=inner_block.paragraph_index,
                    metadata={**inner_block.metadata, "inner_file_hash": inner.file_hash},
                ))
            warnings.extend(f"{info.filename}: {warning}" for warning in inner.warnings)
    return blocks, warnings


def _convert_doc_to_docx(data: bytes, source_file: str) -> bytes:
    """Convert with Word COM in an isolated temporary directory when available."""
    import win32com.client

    with tempfile.TemporaryDirectory(prefix="corporatebrain-doc-") as directory:
        input_path = Path(directory, "input.doc")
        output_path = Path(directory, "output.docx")
        input_path.write_bytes(data)
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = False
        document = None
        try:
            document = word.Documents.Open(str(input_path.resolve()), ReadOnly=True)
            document.SaveAs2(str(output_path.resolve()), FileFormat=16)
        finally:
            if document is not None:
                document.Close(False)
            word.Quit()
        return output_path.read_bytes()


def normalize_document(data: bytes, source_file: str,
                       *, doc_converter: Callable[[bytes, str], bytes] | None = None) -> CanonicalDocument:
    """Normalize original bytes without persistence or indexing side effects."""
    if not isinstance(data, bytes):
        raise TypeError("data must be bytes")
    source_file = Path(source_file).name if "::" not in source_file else source_file
    suffix = Path(source_file.split("::")[-1]).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported document type: {suffix or '<none>'}")
    file_hash = sha256(data).hexdigest()
    warnings: list[str] = []
    blocks: list[CanonicalBlock]
    logical_tables: tuple[LogicalTable, ...] = ()
    if suffix == ".docx":
        logical_tables = read_docx_structure(data)
        blocks = _docx_blocks(data, source_file, file_hash)
    elif suffix == ".doc":
        converter = doc_converter or _convert_doc_to_docx
        try:
            blocks = _docx_blocks(converter(data, source_file), source_file, file_hash)
        except Exception as exc:
            blocks = []
            warnings.append(f"DOC conversion failed: {type(exc).__name__}. No content was extracted.")
    elif suffix == ".pdf":
        blocks, warnings, logical_tables = _pdf_blocks(data, source_file, file_hash)
    elif suffix == ".xlsx":
        blocks = _xlsx_blocks(data, source_file, file_hash)
    elif suffix == ".csv":
        blocks, warnings = _csv_blocks(data, source_file, file_hash)
    else:
        blocks, warnings = _zip_blocks(data, source_file, file_hash)
    if not blocks and not warnings:
        warnings.append("No meaningful content was extracted.")
    return CanonicalDocument(
        document_id=file_hash, file_hash=file_hash, source_file=source_file,
        file_type=suffix.lstrip("."), blocks=tuple(blocks), logical_tables=logical_tables,
        sections=tuple(dict.fromkeys(table.section_path for table in logical_tables if table.section_path)),
        warnings=tuple(warnings),
        metadata={"byte_length": len(data), "normalizer_version": CANONICAL_SCHEMA_VERSION,
                  "logical_table_count": len(logical_tables)},
    )
